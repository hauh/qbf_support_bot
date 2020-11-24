"""Parse dialogue tree from Excel file."""

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from telegram import InlineKeyboardButton, InlineKeyboardMarkup

BACK_BTN = "Назад"
CONTACT_BTN = "Указать контакты"


class ParseError(Exception):
	"""Indicates error in Excel file."""


class Button():
	"""Telegram button from Excel."""

	def __init__(self, row, text, btn_id, back=None):
		self.row = row
		self.text = text
		if text == BACK_BTN:
			self.id = 'back'
		elif text == CONTACT_BTN:
			self.id = 'contact'
		else:
			self.id = btn_id
		self.back = back

	@staticmethod
	def from_cell(cell, back=None):
		text = str(cell.value).lstrip('0123456789')
		if not text or not text.startswith(')'):
			return None
		text = text.removeprefix(')').strip()
		return Button(cell.row, text, hash(text), back)

	def to_telegram(self):
		return [InlineKeyboardButton(self.text, callback_data=self.id)]


def parse_document(file):
	try:
		document = load_workbook(file)
	except OSError as err:
		raise ParseError("Ошибка чтения файла.") from err
	except InvalidFileException as err:
		raise ParseError("Невалидный файл.") from err

	table = document.active
	try:
		menu = {'start': {'message': table[2][0].value, 'buttons': None}}
	except IndexError as err:
		raise ParseError("Пустой файл.") from err

	def build_menu(button, col, max_row):
		try:
			column = next(table.iter_cols(
				min_col=col, min_row=button.row, max_row=max_row))
		except StopIteration:
			return

		if not column[0].value:
			return

		if button.id in menu:
			raise ParseError(f"Дублирующееся меню {button.text}.")

		submenu = menu.setdefault(button.id, {'back': button.back})
		next_buttons = []
		if next_button := Button.from_cell(column[0], button.id):
			next_buttons.append(next_button)
			submenu['message'] = button.text
		else:
			submenu['message'] = column[0].value

		for cell in column[1:]:
			if cell.value:
				if next_button := Button.from_cell(cell, button.id):
					next_buttons.append(next_button)
				else:
					raise ParseError(f"Ошибка в ячейке {cell.coordinate}.")

		if not next_buttons:
			if column[0].value:
				raise ParseError(f"У подменю {button.text} нет кнопок!")
			return

		telegram_buttons = []
		for first_button, second_button in zip(next_buttons, next_buttons[1:]):
			build_menu(first_button, col + 1, second_button.row - 1)
			telegram_buttons.append(first_button.to_telegram())
		build_menu(next_buttons[-1], col + 1, max_row)
		telegram_buttons.append(next_buttons[-1].to_telegram())
		submenu['buttons'] = InlineKeyboardMarkup(telegram_buttons)

	build_menu(Button(3, None, 'main'), table.min_column, table.max_row)
	document.close()
	return menu
