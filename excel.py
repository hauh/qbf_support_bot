"""Parse dialogue tree from Excel file."""

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from telegram import InlineKeyboardMarkup, InlineKeyboardButton


BACK_BTN = "Назад"
CONTACT_BTN = "Указать контакты"


class ParseError(Exception):
	"""Indicates error in Excel file."""


class Button():
	"""Telegram button from Excel."""

	def __init__(self, row, text, btn_id, back=None):
		self.row = row
		self.text = text
		if text == BACK_BTN:
			self.id = 'back'
		elif text == CONTACT_BTN:
			self.id = 'contact'
		else:
			self.id = btn_id
		self.back = back

	@staticmethod
	def from_cell(cell, back=None):
		text = str(cell.value).lstrip('0123456789')
		if not text or not text.startswith(')'):
			return None
		text = text.removeprefix(')').strip()
		return Button(cell.row, text, hash(text), back)

	def to_telegram(self):
		return [InlineKeyboardButton(self.text, callback_data=self.id)]


def parse_document(filename):
	try:
		document = load_workbook(filename=filename)
	except InvalidFileException as err:
		raise ParseError("Невалидный файл.") from err

	table = document.active
	try:
		menu = {'start': table[2][0].value}
	except IndexError as err:
		raise ParseError("Пустой файл.") from err

	def build_menu(button, col, row):
		submenu = menu.setdefault(button.id, {'back': button.back})
		next_buttons = []
		column = next(table.iter_cols(min_col=col, min_row=button.row, max_row=row))
		for cell in column:
			if cell.value:
				if next_button := Button.from_cell(cell, button.id):
					next_buttons.append(next_button)
				elif 'message' not in submenu:
					submenu['message'] = cell.value
				else:
					raise ParseError("Ошибка в ячейке " + cell.coordinate)

		if next_buttons and col != table.max_column:
			for i in range(len(next_buttons) - 1):
				build_menu(next_buttons[i], col + 1, next_buttons[i + 1].row - 1)
			build_menu(next_buttons[-1], col + 1, row)

		if not submenu.get('buttons'):
			submenu['buttons'] = InlineKeyboardMarkup(
				[button.to_telegram() for button in next_buttons])
		submenu.setdefault('message', button.text)

	build_menu(Button(3, None, 'main'), table.min_column, table.max_row)
	document.close()
	return menu
